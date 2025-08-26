#!/usr/bin/env python3
"""
Excel Generator Module - Handles generation of event Excel files
"""

import os
import pandas as pd
import datetime
from typing import List, Dict, Any, Optional, Tuple
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

class EventExcelGenerator:
    """
    Generates Excel files for events with product data, sales tracking and statistics.
    Compatible with the annual event program format.
    """
    
    def __init__(self, db_manager, event_manager):
        """
        Initialize the Event Excel Generator.
        
        Args:
            db_manager: Database manager instance
            event_manager: Event manager instance
        """
        self.db_manager = db_manager
        self.event_manager = event_manager
        self.events_dir = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "data", "events")
        
        # Create events directory if it doesn't exist
        if not os.path.exists(self.events_dir):
            os.makedirs(self.events_dir)
    
    def generate_event_excel(self, event_id: int, output_path: Optional[str] = None) -> str:
        """
        Generate an Excel file for a specific event with products, sales and statistics.
        
        Args:
            event_id: Event ID
            output_path: Optional output path for the Excel file
            
        Returns:
            str: Path to the generated Excel file
        """
        # Get event details
        query = "SELECT name, start_date, end_date, location, description FROM events WHERE id = ?"
        event_details = self.db_manager.execute_query(query, (event_id,))
        
        if not event_details:
            raise ValueError(f"Event with ID {event_id} not found")
        
        event_name = event_details[0][0]
        event_start = event_details[0][1]
        event_location = event_details[0][3]
        event_description = event_details[0][4]
        
        # Get assigned products with categories
        query = """
        SELECT 
            p.id,
            p.sku, 
            p.name, 
            p.description,
            p.category,
            s.name as supplier,
            p.purchase_price,
            COALESCE(ep.event_sale_price, p.sale_price) as sale_price,
            ep.quantity_assigned
        FROM 
            event_products ep
        JOIN 
            products p ON ep.product_id = p.id
        JOIN 
            suppliers s ON p.supplier_id = s.id
        WHERE 
            ep.event_id = ?
        ORDER BY 
            p.category, p.name
        """
        
        assigned_products = self.db_manager.execute_query(query, (event_id,))
        
        # Get sales data
        query = """
        SELECT 
            p.id,
            p.sku,
            SUM(es.quantity) as quantity_sold,
            SUM(es.quantity * es.sale_price) as total_revenue
        FROM 
            event_sales es
        JOIN 
            products p ON es.product_id = p.id
        WHERE 
            es.event_id = ?
        GROUP BY 
            p.id
        """
        
        sales_data = self.db_manager.execute_query(query, (event_id,))
        
        # Create output path if not provided
        if output_path is None:
            safe_event_name = "".join([c if c.isalnum() else "_" for c in event_name])
            output_path = os.path.join(self.events_dir, f"{safe_event_name}_{event_id}.xlsx")
        
        # Convert to DataFrames
        if assigned_products:
            products_df = pd.DataFrame(assigned_products, 
                                      columns=['ID', 'SKU', 'Name', 'Description', 'Category', 
                                              'Supplier', 'Purchase Price', 'Sale Price', 'Quantity Assigned'])
        else:
            products_df = pd.DataFrame(columns=['ID', 'SKU', 'Name', 'Description', 'Category', 
                                              'Supplier', 'Purchase Price', 'Sale Price', 'Quantity Assigned'])
        
        if sales_data:
            sales_df = pd.DataFrame(sales_data, 
                                   columns=['ID', 'SKU', 'Quantity Sold', 'Total Revenue'])
            
            # Merge with products
            merged_df = pd.merge(products_df, sales_df, on=['ID', 'SKU'], how='left')
            merged_df['Quantity Sold'].fillna(0, inplace=True)
            merged_df['Total Revenue'].fillna(0, inplace=True)
            
            # Calculate remaining stock
            merged_df['Remaining'] = merged_df['Quantity Assigned'] - merged_df['Quantity Sold']
            
            # Calculate profit
            merged_df['Profit'] = merged_df['Total Revenue'] - (merged_df['Quantity Sold'] * merged_df['Purchase Price'])
        else:
            merged_df = products_df.copy()
            merged_df['Quantity Sold'] = 0
            merged_df['Total Revenue'] = 0
            merged_df['Remaining'] = merged_df['Quantity Assigned']
            merged_df['Profit'] = 0
        
        # Create Excel writer
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            # Main sheet with products and sales
            self._create_main_sheet(writer, merged_df, event_name, event_start, event_location)
            
            # Summary sheet
            self._create_summary_sheet(writer, merged_df, event_name, event_start, event_location, event_description)
            
            # Category breakdown sheet
            self._create_category_breakdown(writer, merged_df)
            
            # Supplier breakdown sheet
            self._create_supplier_breakdown(writer, merged_df)
            
            # Top products sheet
            if not merged_df.empty and merged_df['Quantity Sold'].sum() > 0:
                self._create_top_products_sheet(writer, merged_df)
        
        # Apply formatting
        self._apply_excel_formatting(output_path)
        
        return output_path
    
    def generate_annual_events_program(self, year: int, output_path: Optional[str] = None) -> str:
        """
        Generate an annual program Excel file with all events and products.
        Format matches the PROGRAMMAPROVVISORIO template.
        
        Args:
            year: Year to generate program for
            output_path: Optional output path for the Excel file
            
        Returns:
            str: Path to the generated Excel file
        """
        # Get all events for the year
        start_date = f"{year}-01-01"
        end_date = f"{year+1}-01-01"
        
        query = """
        SELECT 
            id, 
            name, 
            start_date, 
            end_date, 
            location,
            description
        FROM 
            events
        WHERE 
            start_date >= ? AND start_date < ?
        ORDER BY 
            start_date
        """
        
        events = self.db_manager.execute_query(query, (start_date, end_date))
        
        if not events:
            raise ValueError(f"No events found for year {year}")
        
        # Create DataFrame for events
        events_df = pd.DataFrame(events, columns=['ID', 'Name', 'Start Date', 'End Date', 'Location', 'Description'])
        
        # Get all products that were assigned to any event in this year
        query = """
        SELECT DISTINCT
            p.id,
            p.sku,
            p.name,
            p.description,
            p.category,
            s.name as supplier
        FROM 
            products p
        JOIN 
            event_products ep ON p.id = ep.product_id
        JOIN 
            events e ON ep.event_id = e.id
        JOIN
            suppliers s ON p.supplier_id = s.id
        WHERE 
            e.start_date >= ? AND e.start_date < ?
        ORDER BY
            p.category, p.name
        """
        
        products = self.db_manager.execute_query(query, (start_date, end_date))
        
        if not products:
            products_df = pd.DataFrame(columns=['ID', 'SKU', 'Name', 'Description', 'Category', 'Supplier'])
        else:
            products_df = pd.DataFrame(products, columns=['ID', 'SKU', 'Name', 'Description', 'Category', 'Supplier'])
        
        # Get all sales data for these events
        query = """
        SELECT 
            e.id as event_id,
            e.name as event_name,
            p.id as product_id,
            p.sku,
            SUM(es.quantity) as quantity_sold,
            SUM(es.quantity * es.sale_price) as total_revenue
        FROM 
            event_sales es
        JOIN 
            products p ON es.product_id = p.id
        JOIN 
            events e ON es.event_id = e.id
        WHERE 
            e.start_date >= ? AND e.start_date < ?
        GROUP BY 
            e.id, p.id
        """
        
        sales_data = self.db_manager.execute_query(query, (start_date, end_date))
        
        if not sales_data:
            sales_df = pd.DataFrame(columns=['Event ID', 'Event Name', 'Product ID', 'SKU', 'Quantity Sold', 'Total Revenue'])
        else:
            sales_df = pd.DataFrame(sales_data, 
                                   columns=['Event ID', 'Event Name', 'Product ID', 'SKU', 'Quantity Sold', 'Total Revenue'])
        
        # Create output path if not provided
        if output_path is None:
            output_path = os.path.join(self.events_dir, f"PROGRAMMA_ANNUALE_{year}.xlsx")
        
        # Create Excel writer
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            # Create the main sheet (similar to the template)
            self._create_annual_program_main_sheet(writer, products_df, events_df, sales_df)
            
            # Create the events list sheet
            self._create_events_list_sheet(writer, events_df)
            
            # Create the offer sheet
            self._create_offer_sheet(writer, products_df)
            
            # Create the prizes sheet
            self._create_prizes_sheet(writer, products_df, events_df)
        
        # Apply formatting
        self._apply_annual_program_formatting(output_path)
        
        return output_path
    
    def _create_main_sheet(self, writer, df, event_name, event_date, event_location):
        """Create the main sheet with products and sales data."""
        # Prepare the data
        sheet_data = df[['SKU', 'Name', 'Description', 'Category', 'Supplier', 
                         'Purchase Price', 'Sale Price', 'Quantity Assigned', 
                         'Quantity Sold', 'Remaining', 'Total Revenue', 'Profit']]
        
        # Add header row with event details
        header_df = pd.DataFrame([
            [f"Event: {event_name}", "", "", "", "", "", "", "", "", "", "", ""],
            [f"Date: {event_date}", "", "", "", "", "", "", "", "", "", "", ""],
            [f"Location: {event_location}", "", "", "", "", "", "", "", "", "", "", ""],
            ["", "", "", "", "", "", "", "", "", "", "", ""]
        ])
        
        # Concatenate header and data
        final_df = pd.concat([header_df, sheet_data.reset_index(drop=True)])
        
        # Write to Excel
        final_df.to_excel(writer, sheet_name='Products & Sales', index=False)
    
    def _create_summary_sheet(self, writer, df, event_name, event_date, event_location, event_description):
        """Create a summary sheet with event statistics."""
        # Calculate summary statistics
        total_products = len(df)
        total_assigned = df['Quantity Assigned'].sum()
        total_sold = df['Quantity Sold'].sum()
        total_revenue = df['Total Revenue'].sum()
        total_profit = df['Profit'].sum()
        
        sell_through_rate = (total_sold / total_assigned * 100) if total_assigned > 0 else 0
        
        # Create summary data
        summary_data = {
            'Metric': [
                'Event Name', 'Date', 'Location', 'Description',
                'Total Products', 'Total Assigned', 'Total Sold',
                'Sell-through Rate', 'Total Revenue', 'Total Profit',
                'Average Price', 'Profit Margin'
            ],
            'Value': [
                event_name, event_date, event_location, event_description,
                total_products, total_assigned, total_sold,
                f"{sell_through_rate:.2f}%",
                f"€{total_revenue:.2f}", f"€{total_profit:.2f}",
                f"€{(total_revenue / total_sold if total_sold > 0 else 0):.2f}",
                f"{(total_profit / total_revenue * 100 if total_revenue > 0 else 0):.2f}%"
            ]
        }
        
        summary_df = pd.DataFrame(summary_data)
        summary_df.to_excel(writer, sheet_name='Summary', index=False)
    
    def _create_category_breakdown(self, writer, df):
        """Create a breakdown of sales by product category."""
        if df.empty:
            empty_df = pd.DataFrame(columns=['Category', 'Products', 'Assigned', 'Sold', 'Revenue', 'Profit'])
            empty_df.to_excel(writer, sheet_name='Category Breakdown', index=False)
            return
        
        # Group by category
        category_breakdown = df.groupby('Category').agg({
            'ID': 'count',
            'Quantity Assigned': 'sum',
            'Quantity Sold': 'sum',
            'Total Revenue': 'sum',
            'Profit': 'sum'
        }).reset_index()
        
        # Rename columns
        category_breakdown.columns = ['Category', 'Products', 'Assigned', 'Sold', 'Revenue', 'Profit']
        
        # Add sell-through rate
        category_breakdown['Sell-through Rate'] = category_breakdown.apply(
            lambda row: f"{(row['Sold'] / row['Assigned'] * 100) if row['Assigned'] > 0 else 0:.2f}%", 
            axis=1
        )
        
        # Add totals row
        totals = pd.DataFrame([{
            'Category': 'TOTAL',
            'Products': category_breakdown['Products'].sum(),
            'Assigned': category_breakdown['Assigned'].sum(),
            'Sold': category_breakdown['Sold'].sum(),
            'Revenue': category_breakdown['Revenue'].sum(),
            'Profit': category_breakdown['Profit'].sum(),
            'Sell-through Rate': f"{(category_breakdown['Sold'].sum() / category_breakdown['Assigned'].sum() * 100) if category_breakdown['Assigned'].sum() > 0 else 0:.2f}%"
        }])
        
        category_breakdown = pd.concat([category_breakdown, totals])
        
        # Write to Excel
        category_breakdown.to_excel(writer, sheet_name='Category Breakdown', index=False)
    
    def _create_supplier_breakdown(self, writer, df):
        """Create a breakdown of sales by supplier."""
        if df.empty:
            empty_df = pd.DataFrame(columns=['Supplier', 'Products', 'Assigned', 'Sold', 'Revenue', 'Profit'])
            empty_df.to_excel(writer, sheet_name='Supplier Breakdown', index=False)
            return
        
        # Group by supplier
        supplier_breakdown = df.groupby('Supplier').agg({
            'ID': 'count',
            'Quantity Assigned': 'sum',
            'Quantity Sold': 'sum',
            'Total Revenue': 'sum',
            'Profit': 'sum'
        }).reset_index()
        
        # Rename columns
        supplier_breakdown.columns = ['Supplier', 'Products', 'Assigned', 'Sold', 'Revenue', 'Profit']
        
        # Add sell-through rate
        supplier_breakdown['Sell-through Rate'] = supplier_breakdown.apply(
            lambda row: f"{(row['Sold'] / row['Assigned'] * 100) if row['Assigned'] > 0 else 0:.2f}%", 
            axis=1
        )
        
        # Add totals row
        totals = pd.DataFrame([{
            'Supplier': 'TOTAL',
            'Products': supplier_breakdown['Products'].sum(),
            'Assigned': supplier_breakdown['Assigned'].sum(),
            'Sold': supplier_breakdown['Sold'].sum(),
            'Revenue': supplier_breakdown['Revenue'].sum(),
            'Profit': supplier_breakdown['Profit'].sum(),
            'Sell-through Rate': f"{(supplier_breakdown['Sold'].sum() / supplier_breakdown['Assigned'].sum() * 100) if supplier_breakdown['Assigned'].sum() > 0 else 0:.2f}%"
        }])
        
        supplier_breakdown = pd.concat([supplier_breakdown, totals])
        
        # Write to Excel
        supplier_breakdown.to_excel(writer, sheet_name='Supplier Breakdown', index=False)
    
    def _create_top_products_sheet(self, writer, df):
        """Create a sheet with top-selling products."""
        # Sort by quantity sold
        top_by_quantity = df.sort_values('Quantity Sold', ascending=False).head(10)
        top_by_quantity = top_by_quantity[['SKU', 'Name', 'Category', 'Supplier', 'Quantity Sold', 'Total Revenue', 'Profit']]
        
        # Sort by revenue
        top_by_revenue = df.sort_values('Total Revenue', ascending=False).head(10)
        top_by_revenue = top_by_revenue[['SKU', 'Name', 'Category', 'Supplier', 'Quantity Sold', 'Total Revenue', 'Profit']]
        
        # Write to Excel
        top_by_quantity.to_excel(writer, sheet_name='Top Products', index=False, startrow=1)
        
        # Add header
        worksheet = writer.sheets['Top Products']
        worksheet.cell(row=1, column=1, value="Top Products by Quantity Sold")
    
    def _create_annual_program_main_sheet(self, writer, products_df, events_df, sales_df):
        """Create the main sheet for the annual program."""
        if products_df.empty or events_df.empty:
            # Create empty sheet with basic structure
            pd.DataFrame().to_excel(writer, sheet_name='PROGRAMMA ANNUALE', index=False)
            return
        
        # Create pivot table: Products vs Events with sales quantities
        if not sales_df.empty:
            pivot_df = sales_df.pivot_table(
                values='Quantity Sold',
                index=['Product ID', 'SKU'],
                columns=['Event ID', 'Event Name'],
                aggfunc='sum',
                fill_value=0
            )
            
            # Reset index to get SKU as a column
            pivot_reset = pivot_df.reset_index()
            
            # Merge with product details
            main_df = pd.merge(
                products_df[['ID', 'SKU', 'Name', 'Description', 'Category', 'Supplier']],
                pivot_reset,
                left_on=['ID', 'SKU'],
                right_on=['Product ID', 'SKU'],
                how='left'
            )
            
            # Drop duplicate columns and fill NaN
            main_df = main_df.drop(columns=['Product ID'])
            main_df = main_df.fillna(0)
            
            # Add totals column
            event_columns = [col for col in main_df.columns if isinstance(col, tuple) and len(col) == 2]
            if event_columns:
                main_df['TOTALI'] = main_df[event_columns].sum(axis=1)
        else:
            # No sales data, create basic structure
            main_df = products_df[['ID', 'SKU', 'Name', 'Description', 'Category', 'Supplier']].copy()
            
            # Add empty columns for each event
            for _, event in events_df.iterrows():
                main_df[(event['ID'], event['Name'])] = 0
            
            main_df['TOTALI'] = 0
        
        # Write to Excel
        main_df.to_excel(writer, sheet_name='PROGRAMMA ANNUALE', index=False)
    
    def _create_events_list_sheet(self, writer, events_df):
        """Create a sheet with the list of events."""
        # Prepare the data
        events_list = events_df[['ID', 'Name', 'Start Date', 'End Date', 'Location', 'Description']].copy()
        
        # Add month column
        events_list['Periodo'] = pd.to_datetime(events_list['Start Date']).dt.strftime('%B')
        
        # Reorder columns
        events_list = events_list[['Periodo', 'Start Date', 'End Date', 'Name', 'Location', 'Description', 'ID']]
        
        # Rename columns
        events_list.columns = ['Periodo', 'Data Inizio', 'Data Fine', 'Nome', 'Località', 'Descrizione', 'ID']
        
        # Write to Excel
        events_list.to_excel(writer, sheet_name='GIRO D\'ITALIA', index=False)
    
    def _create_offer_sheet(self, writer, products_df):
        """Create a sheet with product offers."""
        if products_df.empty:
            # Create empty sheet with basic structure
            pd.DataFrame(columns=['Expo', 'EXPO', 'CASCHI', 'OCCHIALI']).to_excel(
                writer, sheet_name='OFFERTA VENDITA', index=False
            )
            return
        
        # Count products by category
        category_counts = products_df['Category'].value_counts().reset_index()
        category_counts.columns = ['Category', 'Count']
        
        # Create offer data
        offer_data = {
            'Expo': ['15/16', ''],
            'EXPO': ['FFWD / CRONO / CST/ BOLLE', ''],
            'CASCHI': [category_counts[category_counts['Category'] == 'Helmets']['Count'].sum() if 'Helmets' in category_counts['Category'].values else 0, ''],
            'OCCHIALI': [category_counts[category_counts['Category'] == 'Sunglasses']['Count'].sum() if 'Sunglasses' in category_counts['Category'].values else 0, '']
        }
        
        offer_df = pd.DataFrame(offer_data)
        
        # Write to Excel
        offer_df.to_excel(writer, sheet_name='OFFERTA VENDITA', index=False)
    
    def _create_prizes_sheet(self, writer, products_df, events_df):
        """Create a sheet for event prizes."""
        # Create basic structure
        prizes_df = pd.DataFrame(columns=['Nome', 'MODELLO', 'MODELLO.1', 'PREZZO ', 'TOTALE'])
        
        # Add event names
        for _, event in events_df.iterrows():
            prizes_df = prizes_df.append({'Nome': event['Name']}, ignore_index=True)
        
        # Write to Excel
        prizes_df.to_excel(writer, sheet_name='ACQUISTO PREMI', index=False)
    
    def _apply_excel_formatting(self, file_path):
        """Apply formatting to the Excel file."""
        # Load workbook
        wb = openpyxl.load_workbook(file_path)
        
        # Define styles
        header_font = Font(bold=True, size=12)
        header_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
        total_font = Font(bold=True)
        total_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        border = Border(
            left=Side(style='thin'), 
            right=Side(style='thin'), 
            top=Side(style='thin'), 
            bottom=Side(style='thin')
        )
        
        # Format each sheet
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            
            # Format header row
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.border = border
            
            # Format data rows
            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    cell.border = border
            
            # Format total rows
            for row in ws.iter_rows():
                if row[0].value == 'TOTAL' or row[0].value == 'TOTALI':
                    for cell in row:
                        cell.font = total_font
                        cell.fill = total_fill
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                adjusted_width = (max_length + 2) * 1.2
                ws.column_dimensions[column_letter].width = min(adjusted_width, 40)
        
        # Save the workbook
        wb.save(file_path)
    
    def _apply_annual_program_formatting(self, file_path):
        """Apply specific formatting to the annual program Excel file."""
        # Load workbook
        wb = openpyxl.load_workbook(file_path)
        
        # Define styles
        header_font = Font(bold=True, size=12)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font_white = Font(bold=True, size=12, color="FFFFFF")
        total_font = Font(bold=True)
        total_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        border = Border(
            left=Side(style='thin'), 
            right=Side(style='thin'), 
            top=Side(style='thin'), 
            bottom=Side(style='thin')
        )
        
        # Format main sheet
        if 'PROGRAMMA ANNUALE' in wb.sheetnames:
            ws = wb['PROGRAMMA ANNUALE']
            
            # Format header row
            for cell in ws[1]:
                cell.font = header_font_white
                cell.fill = header_fill
                cell.border = border
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            # Format data rows
            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    cell.border = border
            
            # Format total column
            total_col = None
            for i, cell in enumerate(ws[1]):
                if cell.value == 'TOTALI':
                    total_col = i + 1
                    break
            
            if total_col:
                for row in ws.iter_rows(min_row=2):
                    cell = row[total_col - 1]
                    cell.font = total_font
                    cell.fill = total_fill
        
        # Format other sheets
        for sheet_name in ['GIRO D\'ITALIA', 'OFFERTA VENDITA', 'ACQUISTO PREMI']:
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                
                # Format header row
                for cell in ws[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.border = border
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # Format data rows
                for row in ws.iter_rows(min_row=2):
                    for cell in row:
                        cell.border = border
        
        # Save the workbook
        wb.save(file_path)
